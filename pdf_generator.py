import csv
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Report:
    """Класс для визуализации статистики
        Attributes:
           profession (string): название профессии
           vacancies_count_by_cities (dict): Словарь - город : количество вакансий
           vacancies_share_by_cities (dict): Словарь - город : процент вакансий от общего кол-ва
           vacancies_count_by_years (dict): Словарь - год : количество вакансий
           vacancies_count_by_years_for_profession (dict): Словарь - год : количество вакансий определённой профессии
           salary_by_years (dict): Словарь - год : уровень зарплат
           salary_by_years_for_profession (dict): Словарь - год : уровень зарплат определённой профессии
           salary_by_cities (dict): Словарь - город : уровень зарплат
           side (Side): стиль стороны ячейки excel
           border (Border): стиль обводки ячейки excel
    """
    def __init__(self, vacancies_count_by_years, vacancies_count_by_years_for_profession, salary_by_years,
                 salary_by_years_for_profession, vacancies_count_by_cities, vacancies_share_by_cities, salary_by_cities,
                 profession):
        """Инициализируект объект Report, формирует различные данные
        Args:
            profession (string): название профессии
            vacancies_count_by_cities (dict): Словарь - город : количество вакансий
            vacancies_share_by_cities (dict): Словарь - город : процент вакансий от общего кол-ва
            vacancies_count_by_years (dict): Словарь - год : количество вакансий
            vacancies_count_by_years_for_profession (dict): Словарь - год : количество вакансий определённой профессии
            salary_by_years (dict): Словарь - год : уровень зарплат
            salary_by_years_for_profession (dict): Словарь - год : уровень зарплат определённой профессии
            salary_by_cities (dict): Словарь - город : уровень зарплат
        """
        self.vacancies_count_by_years = vacancies_count_by_years
        self.vacancies_count_by_years_for_profession = vacancies_count_by_years_for_profession
        self.salary_by_years = salary_by_years
        self.salary_by_years_for_profession = salary_by_years_for_profession
        self.vacancies_count_by_cities = vacancies_count_by_cities
        self.vacancies_share_by_cities = vacancies_share_by_cities
        self.salary_by_cities = salary_by_cities
        self.side = Side(style='thin', color="000000")
        self.thin_border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.profession = profession

    def make_column_stylied(self, ws, index_column, max_row_index):
        """Стилизует ячейки определенной колонки
        Args:
            index_column (int) - индекс колонки
            max_row_index (int) - максимальный индекс  строки
        """
        max_value = -6
        for row in range(1, max_row_index):
            cell = ws.cell(column=index_column, row=row)
            if max_value < len(str(cell.value)):
                max_value = len(str(cell.value))
            cell.border = self.thin_border
        ws.column_dimensions[get_column_letter(index_column)].width = max_value + 2

    def fill_column_by_years(self, ws, column_index, dictionary_values, min_year, max_year):
        """Заполняет колонку в Excel-файле со статистикой по годам
        Args:
            ws (openpyxl.Workbook()): Excel лист
            column_index (int): индекс колонки
            dictionary_values (dict):  словарь - год : значение
            min_year (int): минимальный год для рассмотрения
            max_year (int): максимальный год для рассмотрения
        """
        max_row_index = max_year - min_year
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[value] for value in range(min_year, max_year)]):
            ws.cell(row=row, column=column_index, value=value)

    def fill_column_by_cities(self, ws, column_index, max_row_index, dictionary_values, keys):
        """Заполняет колонку в Excel-файле со статистикой по городам
        Args:
            ws (openpyxl.Workbook()): Excel лист
            column_index (int): индекс колонки
            max_row_index (int): максимальный индекс строки
            dictionary_values (dict):  словарь - город : значение
            keys (list): лист городов
        """
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[key] for key in keys]):
            ws.cell(row=row, column=column_index, value=value)

    def make_ws_by_years(self, ws):
        """Создает excel-страницу со статистикой, составленной по годам
            Args:
                ws (openpyxl.Workbook()): Excel лист
        """
        min_year, max_year = 2007, 2022
        max_column_index = 5
        max_row_index = max_year - min_year
        ws.title = "Статистика по годам"
        self.make_titles(max_column_index,
                         ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
                          f"Количество вакансий - {self.profession}"],
                         ws)
        self.fill_column_by_years(ws, 1,
                                  {x: y for x, y in zip(range(min_year, max_year + 1), range(min_year, max_year + 1))},
                                  min_year, max_year + 1)
        self.fill_column_by_years(ws, 2, self.salary_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 3, self.salary_by_years_for_profession, min_year, max_year + 1)
        self.fill_column_by_years(ws, 4, self.vacancies_count_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 5, self.vacancies_count_by_years_for_profession, min_year, max_year + 1)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 3)

    def make_ws_by_cities(self, ws):
        """Создает excel-страницу со статистикой, составленной по городам
        Args:
            ws (openpyxl.Workbook()): Excel лист
        """
        max_column_index = 5
        max_row_index = 11
        ws.title = "Статистика по городам"
        self.make_titles(max_column_index, ["Город", "Уроввень зарплат", "", "Город", "Доля вакансий"], ws)
        keys = self.salary_by_cities.keys()
        self.fill_column_by_cities(ws, 1, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 2, max_row_index, self.salary_by_cities, keys)
        keys = self.vacancies_share_by_cities.keys()
        self.fill_column_by_cities(ws, 4, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 5, max_row_index, self.vacancies_share_by_cities, keys)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 1)
        ws.column_dimensions[get_column_letter(3)].width = 2
        for cell in ws['E']:
            cell.number_format = '0.00%'

    def make_titles(self, max_column_index, names, ws):
        """Стилизует первую строку в excel
        Args:
            max_column_index (int) - максимальный индекс колонки
            names (string) - значения колонок
            ws (openpyxl.Workbook()): Excel лист
        """
        for cols_index, name in zip([i for i in range(1, max_column_index + 1)], names):
            ws.cell(row=1, column=cols_index, value=name).font = Font(bold=True)

    def generate_excel(self):
        """генерирует excel-файл
        Returns:
            ws_by_cities (openpyxl.Workbook()): Excel лист со статистикой по городам
            ws_by_years (openpyxl.Workbook()): Excel лист со статистикой по годам
        """
        wb = Workbook()
        ws_by_years = wb.active
        ws_by_cities = wb.create_sheet()
        self.make_ws_by_years(ws_by_years)
        self.make_ws_by_cities(ws_by_cities)
        wb.save("report.xlsx")
        return ws_by_years, ws_by_cities

    def generate_normal_bar_graph(self, title, first_line_legend, second_line_legend, fisrt_dict, second_dict, ax):
        """Генерирует график "обычной" гистограммы
            Args:
                ax (axes.SubplotBase) - форма для графика в matplotlib
                title (string) - название графика
                first_line_legend (string) - первая запись в легенде
                second_line_legend (string) - вторая запись в легенде
                first_dict (dict) - словарь для сравнительной характеристики
                second_dict (dict) - словарь для сравнительной характеристики
        """
        labels = [i for i in range(2007, 2023)]
        salary_by_year = [fisrt_dict[key] for key in range(2007, 2023)]
        salary_by_year_for_profession = [second_dict[key] for key in range(2007, 2023)]
        x = np.arange(len(labels))
        width = 0.35
        ax.bar(x - width / 2, salary_by_year, width, label=first_line_legend)
        ax.bar(x + width / 2, salary_by_year_for_profession, width, label=second_line_legend)
        ax.set_title(title)
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(fontsize=8)
        ax.yaxis.grid(True)

    def generate_reverse_bar_graph(self, ax):
        """Генерирует график перевернутой гистограммы
        Args:
            ax (axes.SubplotBase) - форма для графика в matplotlib
        """
        cities = self.salary_by_cities.keys()
        y_pos = np.arange(len(cities))
        performance = [self.salary_by_cities[key] for key in cities]
        ax.barh(y_pos, performance, align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам')
        ax.xaxis.grid(True)

    def generate_pie(self, ax):
        """Генерирует график круговой гистограммы
        Args:
            ax (axes.SubplotBase) - форма для графика в matplotlib
        """
        keys = [*self.vacancies_count_by_cities]
        top_keys, other_keys = keys[:10], keys[10:]
        x = [self.vacancies_count_by_cities[key] for key in top_keys]
        x.append(sum([self.vacancies_count_by_cities[key] for key in other_keys]))
        top_keys.append("Другие")
        ax.pie(x, labels=top_keys)
        ax.set_title('Доля вакансий по городам')

    def generate_image(self):
        """Генерирует картинку"""
        fig = plt.figure()
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        ax1 = fig.add_subplot(221)
        ax2 = fig.add_subplot(222)
        ax3 = fig.add_subplot(223)
        ax4 = fig.add_subplot(224)
        self.generate_normal_bar_graph("Уровень зарплат по годам", "средняя з/п", f"з/п {self.profession}",
                                       self.salary_by_years, self.salary_by_years_for_profession, ax1)
        self.generate_normal_bar_graph("Количество вакансий по годам", "Количество вакансий",
                                       f"Количество вакансий {self.profession}",
                                       self.vacancies_count_by_years, self.vacancies_count_by_years_for_profession, ax2)
        self.generate_reverse_bar_graph(ax3)
        self.generate_pie(ax4)
        plt.tight_layout()
        fig.set_size_inches(9, 7)
        plt.savefig('graph.png')

    def remake_to_percantage(self, ws, column_index):
        """ Выставляет формат ячейки в проценты
        Args:
            ws (openpyxl.Workbook()): Excel лист
            column_index (int): индекс колонки для редактирования
        Returns:
            ws (openpyxl.Workbook()): отформатированный Excel лист
        """
        for row in range(2, ws.max_row + 1):
            ws.cell(column=column_index, row=row).value = str(round((ws.cell(column=column_index, row=row).value * 100), 2)).replace(".", ",") + "%"
        return ws

    def generate_pdf(self):
        """Генерирует пдф-файл
        """
        self.generate_image()
        year_stat, cities_stat = self.generate_excel()
        cities_stat = self.remake_to_percantage(cities_stat, 5)
        image = "graph.png"
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(profession=self.profession, image=image, year_stat=year_stat, cities_stat=cities_stat)
        config = pdfkit.configuration(wkhtmltopdf=r"C:\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={'enable-local-file-access': None})


class Vacancy:
    """Класс для вакансии
        Attributes:
            name (string): название
            salary (string): зарплата
            area_name (string): город
            published_at (int): дата публикации
    """
    def __init__(self, dictionary):
        """Инициализируект объект Vacancy
               Args:
                   dictionary (dict): словарь
        """
        self.currency_to_rub = {"AZN": 35.68,
                                "BYR": 23.91,
                                "EUR": 59.90,
                                "GEL": 21.74,
                                "KGS": 0.76,
                                "KZT": 0.13,
                                "RUR": 1,
                                "UAH": 1.64,
                                "USD": 60.66,
                                "UZS": 0.0055, }
        self.name = dictionary["name"]
        self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * self.currency_to_rub[
            dictionary["salary_currency"]]
        self.area_name = dictionary["area_name"]
        self.published_at = int(dictionary["published_at"][:4])


class DataSet:
    """Составляет базу данных для вакансий.
       Attributes:
           file_name (string): название файла
           profession (string): название профессии
           vacancies_objects (list): список вакансий
           vacancies_count_by_cities (dict): Словарь - город : количество вакансий
           vacancies_share_by_cities (dict): Словарь - город : процент вакансий от общего кол-ва
           vacancies_count_by_years (dict): Словарь - год : количество вакансий
           vacancies_count_by_years_for_profession (dict): Словарь - год : количество вакансий определённой профессии
           salary_by_years (dict): Словарь - год : уровень зарплат
           salary_by_years_for_profession (dict): Словарь - год : уровень зарплат определённой профессии
           salary_by_cities (dict): Словарь - город : уровень зарплат
    """

    def __init__(self, file_name, profession):
        """Инициализируект объект DataSet, создаёт словари данных о профессии
              Args:
                  file_name (string): название файла
                  profession (string): название профессии
        """
        self.file_name = file_name
        self.profession = profession
        csv_read = self.csv_reader()
        dictionaries = self.csv_filer(csv_read[1], csv_read[0])
        vacancies_list = []
        for dictionary in dictionaries:
            vacancies_list.append(Vacancy(dictionary))
        self.vacancies_objects = vacancies_list
        self.vacancies_count_by_cities = self.count_vacancies_by_cities()
        self.vacancies_share_by_cities = self.get_vacancies_share_by_cities()
        self.vacancies_count_by_years = self.count_vacancies_by_years()
        self.vacancies_count_by_years_for_profession = self.count_profession_vacancies_by_years()
        self.salary_by_years = self.get_salary_by_years()
        self.salary_by_years_for_profession = self.get_profession_salary_by_years()
        self.salary_by_cities = self.get_salary_by_cities()

    def check_rows_count(self, rows_count):
        """Принимает количество строк
                  Args:
                      rows_count (int): количество строк"""
        if rows_count == 0:
            print("Пустой файл")
            exit()
        if rows_count == 1:
            print("Нет данных")
            exit()

    def take_ten_items(self, dictionary):
        """Принимает 10 первых пар словарей округляя значения с точностью до 4 знаков после запятой
               Args:
                   dictionary (dict): словарь
               Returns:
                   dict: форматированный словарь"""
        result_dictionary = {}
        for key, i in zip(dictionary, [i for i in range(10)]):
            result_dictionary[key] = round(dictionary[key], 4)
        return result_dictionary

    def csv_reader(self):
        """Создает список вакансий и генерирует списки параметров к ним
              Returns:
                  list: список параметров вакансий
                  list: список вакансий
        """
        vacancies, headlines = [], []
        length, rows_count = 0, 0
        first = True
        with open(self.file_name, encoding="utf-8-sig") as File:
            file = csv.reader(File)
            for row in file:
                rows_count += 1
                if first:
                    length = len(row)
                    headlines = row
                    first = False
                else:
                    is_break = False
                    if length != len(row): is_break = True
                    for word in row:
                        if word == "": is_break = True
                    if is_break: continue
                    vacancies.append(row)
        self.check_rows_count(rows_count)
        return headlines, vacancies

    def csv_filer(self, reader, list_naming):
        """Генерирует словари вакансий и параметров
        Args:
            reader (list): список вакансий
            list_naming (list): список параметров вакансий
        Returns:
            dict: словарь - вакансия : параметры
        """
        dictionaries = []
        for vacancy in reader:
            dictionary = {}
            for name, item in zip(list_naming, vacancy):
                dictionary[name] = item
            dictionaries.append(dictionary)
        return dictionaries

    def get_vacancies_share_by_cities(self):
        """Возвращает словарь городов и процента вакансий от общего кол-ва
            Returns:
                dict: Словарь - город : процент вакансий от общего кол-ва
        """
        dictionary = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                dictionary[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        return self.take_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def get_salary_by_cities(self):
        """Возвращает словарь городов и уровня зарплат
            Returns:
                dict: Словарь - город : уровень зарплат
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_count_by_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + vacancy.salary if vacancy.area_name in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_cities[key])
        return self.take_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def count_vacancies_by_years(self):
        """Возвращает словарь годов и кол-ва вакансий
        Returns:
            dict: Словарь - год : количество вакансий
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary

    def count_profession_vacancies_by_years(self):
        """Возвращает словарь годов и кол-ва вакансий определённой профессии
        Returns:
            dict: Словарь - год : количество вакансий определённой профессии
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def get_salary_by_years(self):
        """Возвращает словарь годов и уровня зарплат
        Returns:
            dict: Словарь - год : уровень зарплат
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.published_at] = (
                dictionary[
                    vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years[key])
        return dict(sorted(dictionary.items(), key=itemgetter(0)))

    def get_profession_salary_by_years(self):
        """Возвращает словарь годов и уровня зарплат определённой профессии
        Returns:
            dict: Словарь - год : уровень зарплат определённой профессии
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[
                    vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years_for_profession[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def count_vacancies_by_cities(self):
        """Возвращает словарь годов и уровня зарплат определённой профессии
        Returns:
            dict: Словарь - год : уровень зарплат определённой профессии
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + 1 if vacancy.area_name in dictionary else 1)
        return dictionary


def get_pdf_statistic():
    """Стартует программу"""
    file_name = input("Введите название файла: ")
    profession = input("Введите название профессии: ")
    dataset = DataSet(file_name, profession)
    Report(dataset.vacancies_count_by_years, dataset.vacancies_count_by_years_for_profession, dataset.salary_by_years,
           dataset.salary_by_years_for_profession,
           dataset.vacancies_count_by_cities, dataset.vacancies_share_by_cities, dataset.salary_by_cities,
           dataset.profession).generate_pdf()
